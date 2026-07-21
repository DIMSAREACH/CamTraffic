"""Parse CSV / XLSX uploads into normalized row dicts."""
from __future__ import annotations

import csv
import io
import re
from typing import Any


def _norm_header(value: str) -> str:
    text = (value or '').strip().lower()
    text = re.sub(r'[\s_\-]+', ' ', text)
    return text


# Canonical field → accepted header aliases (normalized)
HEADER_ALIASES: dict[str, dict[str, str]] = {
    'users': {
        'name': 'name',
        'full name': 'name',
        'fullname name': 'name',
        'email': 'email',
        'phone': 'phone',
        'phone number': 'phone',
        'role': 'role',
        'password': 'password',
    },
    'vehicles': {
        'plate number': 'plate_number',
        'plate': 'plate_number',
        'vehicle type': 'vehicle_type',
        'type': 'vehicle_type',
        'owner email': 'owner_email',
        'owner': 'owner_email',
        'model': 'model',
        'color': 'color',
        'year': 'year',
    },
    'signs': {
        'code': 'code',
        'sign code': 'code',
        'name': 'name',
        'sign name': 'name',
        'category': 'category',
        'description': 'description',
    },
    'cameras': {
        'camera id': 'camera_id',
        'camera code': 'camera_id',
        'code': 'camera_id',
        'location': 'location',
        'name': 'location',
        'road name': 'road_name',
        'road': 'road_name',
        'rtsp url': 'rtsp_url',
        'rtsp': 'rtsp_url',
        'stream url': 'rtsp_url',
        'status': 'status',
    },
    'violations': {
        'plate number': 'plate_number',
        'plate': 'plate_number',
        'violation': 'violation',
        'violation type': 'violation',
        'date': 'date',
        'violation date': 'date',
        'fine': 'fine',
        'fine amount': 'fine',
        'amount': 'fine',
    },
}


def map_headers(import_type: str, raw_headers: list[str]) -> dict[int, str]:
    aliases = HEADER_ALIASES.get(import_type, {})
    mapping: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        key = aliases.get(_norm_header(header))
        if key:
            mapping[idx] = key
    return mapping


def _rows_from_matrix(import_type: str, matrix: list[list[Any]]) -> list[dict[str, str]]:
    if not matrix:
        raise ValueError('File is empty.')
    headers = [str(c or '') for c in matrix[0]]
    mapping = map_headers(import_type, headers)
    if not mapping:
        raise ValueError('No recognized column headers found. Download a template and retry.')

    rows: list[dict[str, str]] = []
    for raw in matrix[1:]:
        if raw is None:
            continue
        # Skip fully empty lines
        if all(cell is None or str(cell).strip() == '' for cell in raw):
            continue
        item: dict[str, str] = {}
        for idx, field in mapping.items():
            cell = raw[idx] if idx < len(raw) else ''
            if cell is None:
                cell = ''
            item[field] = str(cell).strip()
        rows.append(item)
    return rows


def parse_csv_bytes(import_type: str, content: bytes) -> list[dict[str, str]]:
    text = content.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    matrix = [list(row) for row in reader]
    return _rows_from_matrix(import_type, matrix)


def parse_xlsx_bytes(import_type: str, content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('openpyxl is required to import Excel files.') from exc

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    matrix: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append(list(row))
    wb.close()
    return _rows_from_matrix(import_type, matrix)


def parse_upload(import_type: str, file_name: str, content: bytes) -> list[dict[str, str]]:
    name = (file_name or '').lower()
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        return parse_xlsx_bytes(import_type, content)
    if name.endswith('.csv') or name.endswith('.txt'):
        return parse_csv_bytes(import_type, content)
    # sniff
    if content[:2] == b'PK':
        return parse_xlsx_bytes(import_type, content)
    return parse_csv_bytes(import_type, content)
