"""Monthly violations + fines workbook export (openpyxl)."""

from __future__ import annotations

import calendar
from datetime import date, datetime, time
from io import BytesIO

from django.contrib.auth import get_user_model
from django.utils import timezone

User = get_user_model()


def _month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    last_day = calendar.monthrange(year, month)[1]
    start = timezone.make_aware(datetime.combine(date(year, month, 1), time.min))
    end = timezone.make_aware(datetime.combine(date(year, month, last_day), time.max))
    return start, end


def _header_row(ws, headers: list[str]) -> None:
    from openpyxl.styles import Font

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_enforcement_monthly_workbook(*, user, year: int, month: int) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is not installed') from exc

    from fines.models import Fine
    from violations.models import TrafficViolation

    start_dt, end_dt = _month_bounds(year, month)

    violations_qs = (
        TrafficViolation.objects.select_related('driver__user', 'officer__user', 'vehicle')
        .filter(violation_date__gte=start_dt, violation_date__lte=end_dt)
        .order_by('-violation_date')
    )
    fines_qs = (
        Fine.objects.select_related('driver', 'police', 'violation')
        .filter(created_at__gte=start_dt, created_at__lte=end_dt)
        .order_by('-created_at')
    )

    if user.role == 'police':
        violations_qs = violations_qs.filter(officer__user=user)
        fines_qs = fines_qs.filter(police=user)

    violation_fine_ids = {
        f.violation_id: f.id
        for f in fines_qs
        if f.violation_id
    }

    wb = Workbook()
    ws_v = wb.active
    ws_v.title = 'Violations'
    _header_row(ws_v, [
        'ID', 'Date', 'Driver', 'License', 'Type', 'Observed Action',
        'Sign Code', 'Class Key', 'Location', 'Status', 'Plate', 'Officer', 'Fine ID',
    ])
    for v in violations_qs:
        ws_v.append([
            v.id,
            timezone.localtime(v.violation_date).strftime('%Y-%m-%d %H:%M'),
            v.driver.user.full_name if v.driver_id else '',
            v.driver.license_no if v.driver_id else '',
            v.violation_type,
            v.observed_action,
            v.detected_sign_code,
            v.detected_class_key,
            v.location,
            v.status,
            v.vehicle.plate_number if v.vehicle_id else '',
            v.officer.user.full_name if v.officer_id else '',
            violation_fine_ids.get(v.id, ''),
        ])

    ws_f = wb.create_sheet('Fines')
    _header_row(ws_f, [
        'ID', 'Issued', 'Driver', 'License', 'Amount (USD)', 'Reason',
        'Status', 'Location', 'Plate', 'Officer', 'Violation ID', 'Paid At',
    ])
    for f in fines_qs:
        ws_f.append([
            f.id,
            timezone.localtime(f.created_at).strftime('%Y-%m-%d %H:%M'),
            f.driver.full_name if f.driver_id else '',
            f.driver.license_no if f.driver_id else '',
            float(f.amount),
            f.reason,
            f.status,
            f.location,
            f.vehicle_plate,
            f.police.full_name if f.police_id else '',
            f.violation_id or '',
            timezone.localtime(f.paid_at).strftime('%Y-%m-%d %H:%M') if f.paid_at else '',
        ])

    ws_s = wb.create_sheet('Summary')
    ws_s.append(['CamTraffic — Monthly Enforcement Export'])
    ws_s.append(['Year', year])
    ws_s.append(['Month', month])
    ws_s.append(['Violations', violations_qs.count()])
    ws_s.append(['Fines', fines_qs.count()])
    ws_s.append(['Exported by', user.full_name])
    ws_s.append(['Exported at', timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
