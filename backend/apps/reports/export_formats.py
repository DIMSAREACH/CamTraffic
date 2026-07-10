from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


@dataclass
class ReportData:
    title: str
    report_type: str
    generated_at: str
    summary_rows: list[tuple[str, str | int | float]]
    detail_headers: list[str]
    detail_rows: list[list[str | int | float | None]]


def build_csv_content(data: ReportData) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['metric', 'value'])
    for metric, value in data.summary_rows:
        writer.writerow([metric, value])
    writer.writerow([])
    if data.detail_headers:
        writer.writerow(data.detail_headers)
        for row in data.detail_rows:
            writer.writerow(row)
    return output.getvalue()


def build_pdf_bytes(data: ReportData) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title=data.title)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(data.title, styles['Title']),
        Paragraph(f'Generated at: {data.generated_at}', styles['Normal']),
        Spacer(1, 12),
        Paragraph('Summary', styles['Heading2']),
    ]

    summary_table = Table([['Metric', 'Value'], *[[str(a), str(b)] for a, b in data.summary_rows]])
    summary_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )
    elements.append(summary_table)

    if data.detail_headers:
        elements.extend(
            [
                Spacer(1, 16),
                Paragraph('Details', styles['Heading2']),
            ]
        )
        detail_table = Table(
            [data.detail_headers, *[[str(cell) if cell is not None else '' for cell in row] for row in data.detail_rows]],
            repeatRows=1,
        )
        detail_table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ]
            )
        )
        elements.append(detail_table)

    doc.build(elements)
    return buffer.getvalue()


def build_excel_bytes(data: ReportData) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Report'

    bold = Font(bold=True)
    worksheet.append([data.title])
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet.append([f'Generated at: {data.generated_at}'])
    worksheet.append([])
    worksheet.append(['Summary'])
    worksheet['A4'].font = bold
    worksheet.append(['Metric', 'Value'])
    for cell in worksheet[5]:
        cell.font = bold
    for metric, value in data.summary_rows:
        worksheet.append([metric, value])

    if data.detail_headers:
        worksheet.append([])
        worksheet.append(['Details'])
        worksheet.cell(row=worksheet.max_row, column=1).font = bold
        header_row = worksheet.max_row + 1
        worksheet.append(data.detail_headers)
        for cell in worksheet[header_row]:
            cell.font = bold
        for row in data.detail_rows:
            worksheet.append([cell if cell is not None else '' for cell in row])

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
